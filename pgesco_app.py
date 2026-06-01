#!/usr/bin/env python3
"""
PGESCo Engineering Suite — Streamlit Web App
"""

import streamlit as st
import pandas as pd
import numpy as np
import math, re, os, io
from pathlib import Path
from collections import defaultdict
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')

# ══════════════════════════════════════════════════════════════════
#  GOOGLE DRIVE INTEGRATION
# ══════════════════════════════════════════════════════════════════
FOLDER_ID = "1OJHstR9mjF2VNbzQYbtIjnGUKLQgVvjP"

@st.cache_resource
def get_drive_service():
    """Build Google Drive service from Streamlit secrets."""
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        creds = service_account.Credentials.from_service_account_info(
            st.secrets["gcp_service_account"],
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        return build("drive", "v3", credentials=creds)
    except Exception as e:
        return None

def list_drive_files(service, folder_id, mime_filter=None):
    """List files in a Google Drive folder."""
    try:
        q = f"'{folder_id}' in parents and trashed=false"
        if mime_filter:
            q += f" and mimeType='{mime_filter}'"
        results = service.files().list(
            q=q, fields="files(id,name,size,modifiedTime)",
            orderBy="name"
        ).execute()
        return results.get("files", [])
    except Exception as e:
        st.error(f"Drive error: {e}")
        return []

def list_drive_subfolders(service, folder_id):
    """List subfolders in a Drive folder."""
    try:
        q = f"'{folder_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
        results = service.files().list(q=q, fields="files(id,name)").execute()
        return results.get("files", [])
    except Exception:
        return []

def download_drive_file(service, file_id):
    """Download a file from Google Drive as bytes."""
    try:
        from googleapiclient.http import MediaIoBaseDownload
        buf = io.BytesIO()
        request = service.files().get_media(fileId=file_id)
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        buf.seek(0)
        return buf
    except Exception as e:
        st.error(f"Download error: {e}")
        return None

def upload_to_drive(service, folder_id, filename, data_bytes, mime_type):
    """Upload a file to Google Drive folder."""
    try:
        from googleapiclient.http import MediaIoBaseUpload
        from googleapiclient.discovery import build
        meta = {"name": filename, "parents": [folder_id]}
        media = MediaIoBaseUpload(io.BytesIO(data_bytes), mimetype=mime_type)
        f = service.files().create(body=meta, media_body=media, fields="id,name").execute()
        return f
    except Exception as e:
        st.error(f"Upload error: {e}")
        return None

def show_drive_browser(service, folder_id, title="📁 Google Drive Files",
                       file_type=None, multi=False, key="drive"):
    """
    Show a file browser for Google Drive.
    Returns selected file(s) as list of (name, bytes) tuples.
    """
    st.markdown(f"**{title}**")

    # List subfolders
    subfolders = list_drive_subfolders(service, folder_id)
    folder_options = {"📁 Root (PGESCo_Files)": folder_id}
    for sf in subfolders:
        folder_options[f"📁 {sf['name']}"] = sf["id"]

    sel_folder_name = st.selectbox("Folder", list(folder_options.keys()), key=f"{key}_folder")
    active_folder = folder_options[sel_folder_name]

    # List files
    files = list_drive_files(service, active_folder)
    if file_type:
        exts = [file_type] if isinstance(file_type, str) else file_type
        files = [f for f in files if any(f["name"].lower().endswith(e.lower()) for e in exts)]

    if not files:
        st.info("No files found in this folder")
        return []

    file_names = [f["name"] for f in files]

    if multi:
        selected = st.multiselect("Select files", file_names, key=f"{key}_sel")
    else:
        selected = [st.selectbox("Select file", file_names, key=f"{key}_sel")]

    if not selected or (len(selected)==1 and not selected[0]):
        return []

    result = []
    sel_files = [f for f in files if f["name"] in selected]

    if st.button(f"📥 Load selected ({len(sel_files)} file(s))", key=f"{key}_load"):
        prog = st.progress(0)
        for i, f in enumerate(sel_files):
            with st.spinner(f"Downloading {f['name']}…"):
                buf = download_drive_file(service, f["id"])
                if buf:
                    buf.name = f["name"]  # add name attr for compatibility
                    result.append(buf)
            prog.progress((i+1)/len(sel_files))
        prog.empty()
        st.success(f"✓ Loaded {len(result)} file(s)")

    return result


# ══════════════════════════════════════════════════════════════════
#  PAGE CONFIG
# ══════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="PGESCo Engineering Suite",
    page_icon="🔷",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ══════════════════════════════════════════════════════════════════
#  COLOURS & CSS
# ══════════════════════════════════════════════════════════════════
st.markdown("""
<style>
/* ── Base ── */
[data-testid="stAppViewContainer"] {
    background: #0d1117;
    color: #e8eaed;
}
[data-testid="stHeader"] { background: transparent; }
[data-testid="stSidebar"] { background: #161b22; }

/* ── Buttons ── */
.stButton > button {
    background: #00bcd4;
    color: black;
    font-weight: 700;
    border: none;
    border-radius: 6px;
    padding: 10px 24px;
    font-size: 15px;
    width: 100%;
    transition: background 0.2s;
}
.stButton > button:hover { background: #009abc; color: black; }

.btn-orange > button {
    background: #ff7043 !important;
    color: white !important;
}
.btn-orange > button:hover { background: #e64a19 !important; }

.btn-green > button {
    background: #4caf50 !important;
    color: black !important;
}

/* ── Cards ── */
.card {
    background: #1e2738;
    border-radius: 10px;
    padding: 20px;
    margin-bottom: 16px;
    border-left: 4px solid #00bcd4;
}
.card-orange { border-left-color: #ff7043; }

/* ── Metric cards ── */
.metric-card {
    background: #1e2738;
    border-radius: 8px;
    padding: 16px;
    text-align: center;
}
.metric-val { font-size: 2rem; font-weight: 800; color: #00bcd4; }
.metric-val-orange { color: #ff7043; }
.metric-val-blue { color: #5b9cf6; }
.metric-lbl { font-size: 0.8rem; color: #8b949e; }

/* ── Section headers ── */
.section-hdr {
    color: #00e5ff;
    font-size: 0.85rem;
    font-weight: 700;
    letter-spacing: 1px;
    margin-bottom: 8px;
    text-transform: uppercase;
}

/* ── Login card ── */
.login-card {
    background: #1e2738;
    border-radius: 12px;
    padding: 40px;
    max-width: 420px;
    margin: 0 auto;
    border: 1px solid #30363d;
}

/* ── Launcher cards ── */
.launcher-card {
    background: #0e1b2e;
    border-radius: 8px;
    overflow: hidden;
    border: 2px solid #00e5ff;
}
.launcher-card-orange { border-color: #ff8c00; }
.launcher-title {
    background: #00bcd4;
    color: black;
    font-weight: 800;
    font-size: 1rem;
    padding: 12px 16px;
    text-align: center;
    letter-spacing: 1px;
}
.launcher-title-orange { background: #ff8c00; }

/* ── Table ── */
[data-testid="stDataFrame"] { background: #1e2738; }
.stDataFrame { border-radius: 8px; overflow: hidden; }

/* ── Inputs ── */
.stTextInput > div > input, .stSelectbox > div > div {
    background: #1c2430 !important;
    color: #e8eaed !important;
    border-color: #30363d !important;
}

/* ── Divider ── */
hr { border-color: #30363d; }

/* ── Status bar ── */
.status-bar {
    background: #161b22;
    border-radius: 6px;
    padding: 8px 16px;
    font-size: 0.85rem;
    color: #8b949e;
    margin-top: 8px;
}

/* ── Footer ── */
.footer {
    text-align: right;
    color: #4a5568;
    font-size: 0.75rem;
    font-style: italic;
    margin-top: 20px;
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════
#  SESSION STATE INIT
# ══════════════════════════════════════════════════════════════════
def init_state():
    defaults = {
        "logged_in":   False,
        "page":        "login",     # login | launcher | insulation | renamer
        "spec_db":     None,
        "result_rows": [],
        "dwg_results": [],
        "pdf_results": [],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ══════════════════════════════════════════════════════════════════
#  REFERENCE DATA
# ══════════════════════════════════════════════════════════════════
DN_TO_INCH = {15:0.5,20:0.75,25:1.0,32:1.25,40:1.5,50:2.0,65:2.5,80:3.0,
              100:4.0,125:5.0,150:6.0,200:8.0,250:10.0,300:12.0,350:14.0,
              400:16.0,450:18.0,500:20.0,600:24.0,700:28.0,800:32.0,900:36.0,1000:40.0}
INCH_TO_DN = {v:k for k,v in DN_TO_INCH.items()}
PIPE_OD = {15:21.3,20:26.9,25:33.7,40:48.3,50:60.3,80:88.9,100:114.3,
           150:168.3,200:219.1,250:273.0,300:323.9,350:355.6,400:406.4,
           500:508.0,600:610.0,700:711.0,800:813.0,900:914.0,1000:1016.0}
INSULATION_TYPES = {
    "Mineral Wool":      [25,40,50,65,75,80,100],
    "Calcium Silicate":  [25,40,50,65,75,80,100],
    "Rockwool":          [25,40,50,65,75,80,100,120],
    "Polyurethane Foam": [25,40,50,65,75],
    "Cellular Glass":    [25,40,50,65,75,80],
}
VALID_USERS = {"admin":"admin","Admin":"Admin"}

# ══════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════
def parse_inch(s):
    s = str(s).replace('"','').strip()
    m = re.match(r'(\d+)\s+(\d+)/(\d+)',s)
    if m: return int(m.group(1))+int(m.group(2))/int(m.group(3))
    m2 = re.match(r'(\d+)/(\d+)',s)
    if m2: return int(m2.group(1))/int(m2.group(2))
    try: return float(s.split('x')[0].strip())
    except: return 0.0

def inch_to_dn(inch):
    if inch in INCH_TO_DN: return INCH_TO_DN[inch]
    return int(round(inch*25.4/5)*5)

def dn_to_inch(dn):
    if dn in DN_TO_INCH: return DN_TO_INCH[dn]
    keys = sorted(DN_TO_INCH)
    return DN_TO_INCH[min(keys, key=lambda k: abs(k-dn))]

# ══════════════════════════════════════════════════════════════════
#  SPEC DATABASE
# ══════════════════════════════════════════════════════════════════
class SpecDatabase:
    ELBOW_KW  = ("ELBOW",)
    TEE_KW    = ("TEE",)
    FLANGE_KW = ("FLANGE",)
    SKIP = {"Spec Sheet","Branch Table","Spec Data","Spec Data Flag","Sheet",
            "BOLT SET","GASKET","WELDOLET","SOCKOLET","REDUCER","COUPLING",
            "CAP","SWAGE","LEVER","WHEEL","GEAR","VALVE","PRESSURE"}

    def __init__(self):
        self.elbow_df  = pd.DataFrame()
        self.tee_df    = pd.DataFrame()
        self.flange_df = pd.DataFrame()
        self.files_loaded = []

    def load(self, uploaded_files):
        el, te, fl = [], [], []
        for uf in uploaded_files:
            engine = "xlrd" if uf.name.lower().endswith(".xls") else "openpyxl"
            spec_id = re.sub(r'[^A-Za-z0-9]','',Path(uf.name).stem).upper()
            try:
                xl = pd.ExcelFile(uf, engine=engine)
            except Exception:
                continue
            for sheet in xl.sheet_names:
                su = sheet.upper()
                if any(su.startswith(s) for s in self.SKIP): continue
                try:
                    df = pd.read_excel(uf, sheet_name=sheet, header=1, engine=engine)
                except Exception:
                    continue
                df = df.dropna(how="all")
                if df.empty: continue
                if any(su.startswith(k) for k in self.ELBOW_KW):
                    self._elbow(df, sheet, spec_id, el)
                elif any(su.startswith(k) for k in self.TEE_KW):
                    self._tee(df, sheet, spec_id, te)
                elif any(su.startswith(k) for k in self.FLANGE_KW):
                    self._flange(df, sheet, spec_id, fl)
            self.files_loaded.append(uf.name)
        self.elbow_df  = pd.DataFrame(el) if el else pd.DataFrame()
        self.tee_df    = pd.DataFrame(te) if te else pd.DataFrame()
        self.flange_df = pd.DataFrame(fl) if fl else pd.DataFrame()

    def _col(self, df, cands):
        for c in cands:
            if c in df.columns: return c
            for col in df.columns:
                if str(col).strip().lower() == c.lower(): return col
        return None

    def _elbow(self, df, sheet, spec_id, out):
        su = sheet.upper()
        angle = 45 if "45" in su else 90
        cs, cr = self._col(df,["Sizes","Size"]), self._col(df,["R"])
        if not cs or not cr: return
        for _,row in df.iterrows():
            sz = str(row[cs]).strip() if not pd.isna(row.get(cs,float('nan'))) else ""
            rv = row.get(cr)
            if not sz or sz in("Sizes","nan","Size") or pd.isna(rv): continue
            try:
                rv = float(rv)
                inch = parse_inch(sz)
                if inch>0 and rv>0:
                    out.append({"dn":inch_to_dn(inch),"r_mm":round(rv*25.4,2),"angle":angle})
            except: continue

    def _tee(self, df, sheet, spec_id, out):
        cs = self._col(df,["Sizes","Size"])
        l1,l2,l3 = self._col(df,["L1"]),self._col(df,["L2"]),self._col(df,["L3"])
        if not cs or not l1 or not l2 or not l3: return
        for _,row in df.iterrows():
            sz = str(row[cs]).strip() if not pd.isna(row.get(cs,float('nan'))) else ""
            if not sz or sz in("Sizes","nan","Size"): continue
            try:
                v1,v2,v3 = float(row[l1]),float(row[l2]),float(row[l3])
                parts = re.split(r'[xX]',sz.replace('"',''))
                mi = parse_inch(parts[0]) if parts else 0
                bi = parse_inch(parts[1]) if len(parts)>1 else mi
                if mi>0:
                    out.append({"dn":inch_to_dn(mi),"dn_branch":inch_to_dn(bi),
                                "l1_mm":round(v1*25.4,2),"l2_mm":round(v2*25.4,2),
                                "l3_mm":round(v3*25.4,2)})
            except: continue

    def _flange(self, df, sheet, spec_id, out):
        cs = self._col(df,["Sizes","Size"])
        cl = self._col(df,["L"])
        cd1= self._col(df,["D1"])
        cnd= self._col(df,["Nominal Diameter_S1","Nominal Diameter_S-ALL"])
        if not cs or not cl: return
        for _,row in df.iterrows():
            sz = str(row[cs]).strip() if not pd.isna(row.get(cs,float('nan'))) else ""
            lv = row.get(cl)
            if not sz or sz in("Sizes","nan","Size") or pd.isna(lv): continue
            try:
                lv = float(lv)
                if lv<=0: continue
                d1 = float(row[cd1]) if cd1 and not pd.isna(row.get(cd1,float('nan'))) else 0
                nd = float(row[cnd]) if cnd and not pd.isna(row.get(cnd,float('nan'))) else 0
                inch = parse_inch(sz)
                if inch<=0: inch = nd if nd>0 else 0
                if inch<=0: continue
                out.append({"dn":inch_to_dn(inch),"l_mm":round(lv*25.4,2),"d1_in":d1,"nd_in":nd if nd>0 else inch})
            except: continue

    def find_elbow(self, dn, angle=90):
        if self.elbow_df.empty: return None
        m = self.elbow_df[self.elbow_df.dn==dn]
        if m.empty:
            m = self.elbow_df.iloc[(self.elbow_df.dn-dn).abs().argsort()[:1]]
        if m.empty: return None
        ma = m[m.angle==angle]
        return float((ma if not ma.empty else m).iloc[0]["r_mm"])

    def find_tee(self, dn, dn_b):
        if self.tee_df.empty: return None
        m = self.tee_df[(self.tee_df.dn==dn)&(self.tee_df.dn_branch==dn_b)]
        if m.empty: m = self.tee_df[self.tee_df.dn==dn]
        if m.empty: m = self.tee_df.iloc[(self.tee_df.dn-dn).abs().argsort()[:1]]
        return m.iloc[0].to_dict() if not m.empty else None

    def find_flange(self, dn):
        if self.flange_df.empty: return None
        m = self.flange_df[self.flange_df.dn==dn]
        if m.empty: m = self.flange_df.iloc[(self.flange_df.dn-dn).abs().argsort()[:1]]
        return m.iloc[0].to_dict() if not m.empty else None

# ══════════════════════════════════════════════════════════════════
#  MTO PARSER
# ══════════════════════════════════════════════════════════════════
class MTOParser:
    TYPES = {"PIPE":["PIPE"],"ELBOW":["ELBOW"],"TEE":["TEE"],
             "FLANGE":["FLANGE","BLIND"],"NOZZLE":["NOZZLE","OLET","WELDOLET"]}

    def parse(self, uploaded_file):
        engine = "xlrd" if uploaded_file.name.lower().endswith(".xls") else "openpyxl"
        df = pd.read_excel(uploaded_file, engine=engine, header=None)
        hr = self._find_header(df)
        if hr is None:
            raise ValueError("Header row not found — need QTY and SIZE columns")
        hdr = df.iloc[hr]
        c_thk=c_spec=c_size=c_desc=c_qty=c_unit=None
        for ci,v in enumerate(hdr):
            vs = str(v).strip().upper() if not pd.isna(v) else ""
            if "INSULATION" in vs and "THICK" in vs: c_thk=ci
            elif "PIPE" in vs and "CLASS" in vs:     c_spec=ci
            elif "SIZE" in vs or vs=="DN":            c_size=ci
            elif "DESCRIPTION" in vs or vs in("DES.","DES"): c_desc=ci
            elif vs in("QTY","QUANTITY"):             c_qty=ci
            elif vs=="UNIT":                          c_unit=ci
        if not c_desc or not c_qty or not c_size:
            raise ValueError("Cannot find DESCRIPTION/QTY/SIZE columns")
        items=[]; cur_spec=""; cur_thk=None
        for ri in range(hr+1,len(df)):
            row=df.iloc[ri]
            if c_thk and not pd.isna(row.iloc[c_thk]):
                try: cur_thk=int(float(row.iloc[c_thk]))
                except: pass
            if c_spec and not pd.isna(row.iloc[c_spec]):
                s=str(row.iloc[c_spec]).strip()
                if s not in("PIPE CLASS",""): cur_spec=s
            desc=row.iloc[c_desc]; size=row.iloc[c_size]; qty=row.iloc[c_qty]
            unit=str(row.iloc[c_unit]).strip().lower() if c_unit and not pd.isna(row.iloc[c_unit]) else ""
            if pd.isna(desc) or pd.isna(qty): continue
            desc_s=str(desc).strip()
            if not desc_s or desc_s.upper() in("DESCRIPTION","DES.","DES"): continue
            try: qty_f=float(qty)
            except: continue
            if qty_f<=0: continue
            dn=0
            if not pd.isna(size):
                try: dn=int(float(str(size).split("x")[0].strip()))
                except: pass
            items.append({"ins_thickness":cur_thk,"spec":cur_spec,"dn":dn,
                          "description":desc_s,"quantity":qty_f,"unit":unit,
                          "type":self._classify(desc_s)})
        return items

    def _find_header(self, df):
        for i in range(min(30,len(df))):
            vals=[str(v).strip().upper() for v in df.iloc[i] if not pd.isna(v)]
            combined=" ".join(vals)
            if ("SIZE" in combined or "DN" in vals) and ("QTY" in vals or "QUANTITY" in combined):
                return i
            if "INSULATION THICKNESS" in combined or "PIPE CLASS" in combined:
                for j in range(i,min(i+4,len(df))):
                    v2=[str(v).strip().upper() for v in df.iloc[j] if not pd.isna(v)]
                    if "QTY" in v2 or "QUANTITY" in " ".join(v2): return j
        return None

    def _classify(self, desc):
        du=desc.upper()
        for t,kws in self.TYPES.items():
            if any(k in du for k in kws): return t
        return "OTHER"

# ══════════════════════════════════════════════════════════════════
#  INSULATION CALCULATOR
# ══════════════════════════════════════════════════════════════════
class InsulationCalculator:
    def __init__(self, db): self.db=db

    def calculate(self, items, ins_type, tmm):
        rows=[]; warns=[]; rid=1
        for item in items:
            if item["type"] in("NOZZLE","OTHER"): continue
            try:
                nr,nw=self._process(item,ins_type,tmm,rid)
                rows.extend(nr); warns.extend(nw); rid+=len(nr)
            except Exception as e:
                warns.append(f"Error '{item['description'][:40]}': {e}")
        return rows,warns

    def consolidate(self, rows, ins_type, tmm):
        groups={}
        for r in rows:
            key=(r["Spec"],r["_dn"],r["_type"])
            if key not in groups:
                groups[key]={"Spec":r["Spec"],"Size":r["Size"],
                             "Description":r["Description"],"Length (m)":0.0,
                             "_dn":r["_dn"],"_type":r["_type"]}
            groups[key]["Length (m)"]+=r["Length (m)"]
        out=[]
        for i,row in enumerate(sorted(groups.values(),key=lambda x:(x["_dn"],x["_type"])),1):
            row["ID"]=i; row["Length (m)"]=round(row["Length (m)"],3); out.append(row)
        return out

    def _process(self, item, ins_type, tmm, sid):
        itype=item["type"]; desc=item["description"]; dn=item["dn"]
        qty=item["quantity"]; spec=item["spec"]; unit=item.get("unit","")
        ins_desc=f"{ins_type} {tmm}mm"; rows=[]; warns=[]
        if itype=="PIPE":
            lm=round(qty/1000,4) if (unit=="mm" or qty>500) else round(qty,4)
            rows.append(self._row(sid,spec,dn,ins_desc,lm,"Pipe"))
        elif itype=="ELBOW":
            angle=45 if "45" in desc else 90
            r_mm=self.db.find_elbow(dn,angle)
            if not r_mm:
                warns.append(f"Elbow DN{dn} not in spec — estimated")
                r_mm=PIPE_OD.get(dn,dn*1.05)*1.5
            lm=round(r_mm*qty/1000,4)
            rows.append(self._row(sid,spec,dn,ins_desc,lm,"Elbow"))
        elif itype=="TEE":
            dn_b=dn
            m=re.search(r'(\d+)\s*[xX]\s*(\d+)',desc)
            if m: dn_b=int(m.group(2))
            tr=self.db.find_tee(dn,dn_b)
            if tr:
                l1,l2,l3=tr["l1_mm"],tr["l2_mm"],tr["l3_mm"]
            else:
                warns.append(f"Tee DN{dn} not in spec — estimated")
                od=PIPE_OD.get(dn,dn*1.05)
                l1=l2=od*0.75; l3=PIPE_OD.get(dn_b,dn_b*1.05)*0.5
            lm_main=round((l1+l2)*qty/1000,4)
            rows.append(self._row(sid,spec,dn,ins_desc,lm_main,"Tee"))
            lm_b=round(l3*qty/1000,4)
            if dn_b!=dn: rows.append(self._row(sid+1,spec,dn_b,ins_desc,lm_b,"Tee"))
            else: rows[-1]["Length (m)"]=round(rows[-1]["Length (m)"]+lm_b,4)
        elif itype=="FLANGE":
            fr=self.db.find_flange(dn)
            if fr:
                l_mm=fr["l_mm"]; d1=fr.get("d1_in",0); nd=fr.get("nd_in",dn_to_inch(dn))
                cf=d1/nd if nd>0 and d1>0 else 1.0
            else:
                warns.append(f"Flange DN{dn} not in spec — estimated")
                l_mm=PIPE_OD.get(dn,dn*1.05)*0.5; cf=1.0
            lm=round(l_mm*cf*qty/1000,4)
            rows.append(self._row(sid,spec,dn,ins_desc,lm,"Flange"))
        return rows,warns

    @staticmethod
    def _row(rid,spec,dn,ins_desc,lm,ftype):
        return {"ID":rid,"Spec":spec,"Size":f"DN{dn}","Description":ins_desc,
                "Length (m)":round(lm,4),"_dn":dn,"_type":ftype}

# ══════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════
def export_to_excel(rows):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill,Font,Alignment,Border,Side
    from openpyxl.utils import get_column_letter
    wb=Workbook(); ws=wb.active; ws.title="Insulation MTO"
    thin=Side(style="thin",color="30363d")
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    def hc(ws,r,c,v,fg="E8EAD0",bg="1C2430",bold=True,sz=10):
        cell=ws.cell(r,c,v)
        cell.fill=PatternFill("solid",fgColor=bg)
        cell.font=Font(bold=bold,color=fg,size=sz)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=bdr
        return cell
    def dc(ws,r,c,v,fg="E8EAD0",bg="1E2738"):
        cell=ws.cell(r,c,v)
        cell.fill=PatternFill("solid",fgColor=bg)
        cell.font=Font(color=fg,size=9)
        cell.alignment=Alignment(horizontal="center",vertical="center")
        cell.border=bdr
        return cell
    ws.merge_cells("A1:E1")
    ws["A1"].value="PGESCo — Insulation MTO Report"
    ws["A1"].font=Font(bold=True,size=14,color="FF7043")
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws["A1"].fill=PatternFill("solid",fgColor="161B22")
    ws.row_dimensions[1].height=30
    ws.merge_cells("A2:E2")
    ws["A2"].value=f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M')}"
    ws["A2"].font=Font(size=9,color="8B949E")
    ws["A2"].alignment=Alignment(horizontal="center")
    ws["A2"].fill=PatternFill("solid",fgColor="161B22")
    ws.row_dimensions[3].height=4
    for ci,h in enumerate(["ID","Spec","Size (DN)","Description","Length (m)"],1):
        hc(ws,4,ci,h,fg="00E5FF")
    ws.row_dimensions[4].height=28
    for ri,row in enumerate(rows,5):
        dn=row.get("_dn",0)
        bg="1E2738" if ri%2==0 else "161B22"
        fg="FF7043" if dn>300 else "E8EAD0"
        dc(ws,ri,1,row.get("ID",ri-4),bg=bg)
        dc(ws,ri,2,row.get("Spec",""),bg=bg)
        dc(ws,ri,3,row.get("Size",f"DN{dn}"),fg=fg,bg=bg)
        dc(ws,ri,4,row.get("Description",""),bg=bg)
        dc(ws,ri,5,row.get("Length (m)",0),bg=bg)
        ws.row_dimensions[ri].height=20
    tr=len(rows)+5
    ws.merge_cells(f"A{tr}:D{tr}")
    hc(ws,tr,1,"TOTAL INSULATION LENGTH",bg="00BCD4",fg="000000")
    hc(ws,tr,5,f"{sum(r.get('Length (m)',0) for r in rows):.3f} m",bg="1C2430",fg="00E5FF")
    for ci,w in enumerate([8,12,12,34,14],1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════
#  PAGES
# ══════════════════════════════════════════════════════════════════
def page_login():
    st.markdown("<br><br>", unsafe_allow_html=True)
    _,col,_ = st.columns([1,1.4,1])
    with col:
        st.markdown("""
        <div style='text-align:center; margin-bottom:24px;'>
            <div style='font-size:2.2rem; font-weight:900; color:#1565C0; letter-spacing:2px;'>PGESCO</div>
            <div style='font-size:0.8rem; color:#8b949e;'>Integrated Innovative Engineering</div>
            <div style='height:2px; background:#00bcd4; margin:14px 0;'></div>
            <div style='font-size:0.85rem; font-weight:700; color:#00e5ff; letter-spacing:1px;'>PGESCO ENGINEERING SUITE</div>
        </div>
        """, unsafe_allow_html=True)

        with st.container():
            st.markdown("<div style='background:#1e2738;padding:32px;border-radius:12px;border:1px solid #30363d;'>", unsafe_allow_html=True)
            st.markdown("<p style='color:#e8eaed;font-size:1.3rem;font-weight:700;text-align:center;margin-bottom:4px;'>Welcome Back</p>", unsafe_allow_html=True)
            st.markdown("<p style='color:#8b949e;font-size:0.85rem;text-align:center;margin-bottom:20px;'>Sign in to PGESCo Engineering Suite</p>", unsafe_allow_html=True)

            username = st.text_input("USERNAME", placeholder="Enter username", label_visibility="visible")
            password = st.text_input("PASSWORD", placeholder="Enter password", type="password")

            if st.button("SIGN IN  →", use_container_width=True):
                if VALID_USERS.get(username) == password or (username.lower()=="admin" and password.lower()=="admin"):
                    st.session_state.logged_in = True
                    st.session_state.page = "launcher"
                    st.rerun()
                else:
                    st.error("✗  Invalid username or password")
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("""
        <div class='footer' style='text-align:center; margin-top:16px;'>
            © 2025 PGESCo — Integrated Innovative Engineering<br>
            <i>Designed by / Eng. Shady Ahmed Elkenany</i>
        </div>
        """, unsafe_allow_html=True)


def page_launcher():
    # Header
    st.markdown("""
    <div style='background:#161b22;padding:14px 24px;margin:-1rem -1rem 0 -1rem;border-bottom:2px solid #00bcd4;display:flex;align-items:center;justify-content:space-between;'>
        <div>
            <span style='font-size:1.5rem;font-weight:900;color:#1565C0;'>PGESCO</span>
            <span style='font-size:0.7rem;color:#8b949e;margin-left:8px;'>Integrated Innovative Engineering</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<h2 style='text-align:center;color:#e8eaed;font-weight:900;letter-spacing:2px;'>WELCOME TO PGESCO ENGINEERING SUITE</h2>", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    col1, col2 = st.columns(2, gap="large")

    with col1:
        st.markdown("""
        <div style='background:#0e1b2e;border:2px solid #00e5ff;border-radius:8px;overflow:hidden;'>
            <div style='background:#00bcd4;padding:12px;text-align:center;font-weight:900;font-size:1rem;color:black;letter-spacing:1px;'>
                📊  INSULATION MTO ANALYTICS
            </div>
            <div style='padding:16px;color:#a0aec0;font-size:0.9rem;text-align:center;'>
                Analyze piping MTO files for insulation properties,
                calculate lengths for PIPE, ELBOW, TEE & FLANGE
                from spec sheets. Export results directly to Excel.
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Open Insulation MTO  →", key="btn_ins", use_container_width=True):
            st.session_state.page = "insulation"
            st.rerun()

    with col2:
        st.markdown("""
        <div style='background:#0e1b2e;border:2px solid #ff8c00;border-radius:8px;overflow:hidden;'>
            <div style='background:#ff8c00;padding:12px;text-align:center;font-weight:900;font-size:1rem;color:black;letter-spacing:1px;'>
                🗂️  FILE RENAMER & PDF TOOLS
            </div>
            <div style='padding:16px;color:#a0aec0;font-size:0.9rem;text-align:center;'>
                Batch-rename DWG files using Line Number lookup
                from Excel. Merge multi-sheet PDFs per line number,
                auto-convert portrait pages to landscape.
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="btn-orange">', unsafe_allow_html=True)
        if st.button("Open File Renamer  →", key="btn_ren", use_container_width=True):
            st.session_state.page = "renamer"
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("<br><br>")
    st.markdown("<div class='footer'>Please select an application to proceed. &nbsp;|&nbsp; <i>Designed by / Eng. Shady Ahmed Elkenany</i></div>", unsafe_allow_html=True)


def page_insulation():
    # ── Header ───────────────────────────────────────────────────
    hcol1, hcol2 = st.columns([1,4])
    with hcol1:
        if st.button("← Suite Home"):
            st.session_state.page = "launcher"; st.rerun()
    with hcol2:
        st.markdown("<h2 style='color:#e8eaed;margin:0;'>Insulation MTO Analytics</h2>", unsafe_allow_html=True)
    st.markdown("<div style='height:2px;background:#00bcd4;margin-bottom:16px;'></div>", unsafe_allow_html=True)

    # ── Layout ───────────────────────────────────────────────────
    left, right = st.columns([1, 2], gap="medium")

    with left:
        # Spec files
        st.markdown("<div class='section-hdr'>① Spec Files (C18, C38…)</div>", unsafe_allow_html=True)
        drive = get_drive_service()
        if drive is None:
            st.warning("Google Drive not connected")
            spec_files = st.file_uploader("Or upload directly", type=["xlsx","xls"], accept_multiple_files=True, label_visibility="collapsed")
        else:
            spec_files = show_drive_browser(drive, FOLDER_ID, "📁 Select Spec Files", file_type=[".xlsx",".xls"], multi=True, key="spec")
        if spec_files:
            if st.button("⚡ Load Specs", use_container_width=True):
                db = SpecDatabase()
                with st.spinner("Reading spec files…"):
                    db.load(spec_files)
                st.session_state.spec_db = db
                st.success(f"✓ ELBOW:{len(db.elbow_df)}  TEE:{len(db.tee_df)}  FLANGE:{len(db.flange_df)}")

        st.markdown("---")

        # Insulation settings
        st.markdown("<div class='section-hdr'>② Insulation Settings</div>", unsafe_allow_html=True)
        ins_type = st.selectbox("Type", list(INSULATION_TYPES.keys()))
        thickness = st.selectbox("Thickness (mm)", INSULATION_TYPES[ins_type])

        st.markdown("---")

        # MTO file
        st.markdown("<div class='section-hdr'>③ Raw Piping MTO</div>", unsafe_allow_html=True)
        if drive is None:
            mto_files = st.file_uploader("Or upload directly", type=["xlsx","xls"], label_visibility="collapsed")
            mto_file = mto_files
        else:
            mto_loaded = show_drive_browser(drive, FOLDER_ID, "📁 Select MTO File", file_type=[".xlsx",".xls"], multi=False, key="mto")
            mto_file = mto_loaded[0] if mto_loaded else None
        if mto_file:
            st.success(f"✓ {mto_file.name}")

        st.markdown("<br>", unsafe_allow_html=True)
        process_clicked = st.button("⚙   PROCESS MTO", use_container_width=True, type="primary")

    with right:
        db = st.session_state.get("spec_db")

        if process_clicked:
            if not mto_file:
                st.warning("Please upload a Raw Piping MTO file first")
            else:
                with st.spinner("Processing MTO…"):
                    try:
                        parser = MTOParser()
                        items  = parser.parse(mto_file)
                        calc   = InsulationCalculator(db or SpecDatabase())
                        rows, warns = calc.calculate(items, ins_type, thickness)
                        consolidated = calc.consolidate(rows, ins_type, thickness)
                        st.session_state.result_rows = consolidated
                        if warns:
                            with st.expander(f"⚠ {len(warns)} spec warnings"):
                                for w in warns: st.write(f"- {w}")
                    except Exception as e:
                        st.error(f"Error: {e}")

        result_rows = st.session_state.get("result_rows", [])

        if result_rows:
            # Stats
            total_len  = sum(r.get("Length (m)",0) for r in result_rows)
            unique_dns = len(set(r.get("_dn",0) for r in result_rows))
            c1,c2,c3 = st.columns(3)
            with c1:
                st.markdown(f"<div class='metric-card'><div class='metric-lbl'>Total Insulation Length</div><div class='metric-val'>{total_len:,.3f}</div><div class='metric-lbl'>m</div></div>", unsafe_allow_html=True)
            with c2:
                st.markdown(f"<div class='metric-card'><div class='metric-lbl'>Total Line Items</div><div class='metric-val metric-val-blue'>{len(result_rows)}</div><div class='metric-lbl'>items</div></div>", unsafe_allow_html=True)
            with c3:
                st.markdown(f"<div class='metric-card'><div class='metric-lbl'>Unique Pipe Sizes</div><div class='metric-val metric-val-orange'>{unique_dns}</div><div class='metric-lbl'>DNs</div></div>", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # Chart
            dn_map = defaultdict(float)
            for r in result_rows:
                dn_map[r["_dn"]] += r.get("Length (m)",0)
            if dn_map:
                fig, ax = plt.subplots(figsize=(5,3), facecolor="#161b22")
                ax.set_facecolor("#161b22")
                colors = ["#00e5ff","#ff7043","#5b9cf6","#4caf50","#ffd700","#e040fb","#ff6090","#00bfa5","#ffab40"]
                dns = sorted(dn_map)
                vals = [dn_map[d] for d in dns]
                labels = [f"DN{d}" for d in dns]
                wedges,_,atexts = ax.pie(vals, colors=colors[:len(dns)],
                                          autopct="%1.1f%%", pctdistance=0.82,
                                          wedgeprops=dict(width=0.52, edgecolor="#161b22", linewidth=2),
                                          startangle=90)
                for at in atexts: at.set_color("#0d1117"); at.set_fontsize(7)
                ax.legend(wedges, labels, loc="lower center", bbox_to_anchor=(0.5,-0.22),
                          ncol=min(len(dns),6), fontsize=7, framealpha=0, labelcolor="white")
                fig.tight_layout(pad=0.4)
                st.pyplot(fig)
                plt.close(fig)

            st.markdown("<br>", unsafe_allow_html=True)

            # Table
            st.markdown("<div class='section-hdr'>Insulation MTO — Preview</div>", unsafe_allow_html=True)
            df_show = pd.DataFrame([{
                "ID": r["ID"], "Spec": r["Spec"], "Size": r["Size"],
                "Description": r["Description"], "Length (m)": r["Length (m)"]
            } for r in result_rows])
            st.dataframe(df_show, use_container_width=True, hide_index=True)

            # Export
            excel_bytes = export_to_excel(result_rows)
            fname_out = f"Insulation_MTO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(
                label="⬇  Download Insulation MTO (Excel)",
                data=excel_bytes,
                file_name=fname_out,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            drive_up = get_drive_service()
            if drive_up and st.button("☁  Save to Google Drive", key="save_drive", use_container_width=True):
                with st.spinner("Uploading to Drive…"):
                    res = upload_to_drive(drive_up, FOLDER_ID, fname_out, excel_bytes,
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    if res:
                        st.success(f"✓ Saved to Drive: {res['name']}")
        else:
            st.info("Upload spec files, set insulation type, upload MTO file, then click PROCESS MTO")

    st.markdown("<div class='footer'>Designed by / Eng. Shady Ahmed Elkenany</div>", unsafe_allow_html=True)


def page_renamer():
    hcol1, hcol2 = st.columns([1,4])
    with hcol1:
        if st.button("← Suite Home"):
            st.session_state.page = "launcher"; st.rerun()
    with hcol2:
        st.markdown("<h2 style='color:#e8eaed;margin:0;'>File Renamer & PDF Tools</h2>", unsafe_allow_html=True)
    st.markdown("<div style='height:2px;background:#ff7043;margin-bottom:16px;'></div>", unsafe_allow_html=True)

    left, right = st.columns([1, 2], gap="medium")

    with left:
        st.markdown("<div class='section-hdr'>① Source Files</div>", unsafe_allow_html=True)
        drive_r = get_drive_service()
        if drive_r is None:
            excel_file = st.file_uploader("Excel Sheet (Line Numbers)", type=["xlsx","xls"])
            dwg_files  = st.file_uploader("DWG Files", type=["dwg"], accept_multiple_files=True)
            pdf_files  = st.file_uploader("PDF Files", type=["pdf"], accept_multiple_files=True)
        else:
            ex_loaded = show_drive_browser(drive_r, FOLDER_ID, "📁 Select Excel Sheet", file_type=[".xlsx",".xls"], multi=False, key="ren_ex")
            excel_file = ex_loaded[0] if ex_loaded else None
            dwg_loaded = show_drive_browser(drive_r, FOLDER_ID, "📁 Select DWG Files", file_type=[".dwg"], multi=True, key="ren_dwg")
            dwg_files  = dwg_loaded if dwg_loaded else []
            pdf_loaded = show_drive_browser(drive_r, FOLDER_ID, "📁 Select PDF Files", file_type=[".pdf"], multi=True, key="ren_pdf")
            pdf_files  = pdf_loaded if pdf_loaded else []

        st.markdown("---")
        st.markdown("<div class='section-hdr'>② DWG Settings</div>", unsafe_allow_html=True)
        dwg_suffix = st.text_input("DWG Suffix", placeholder="_R0")

        st.markdown("<div class='section-hdr'>③ PDF Settings</div>", unsafe_allow_html=True)
        pdf_suffix = st.text_input("PDF Suffix", placeholder="_IFC")

        st.markdown("---")
        st.markdown("<div class='section-hdr'>④ Actions</div>", unsafe_allow_html=True)
        run_dwg = st.button("🗂  Rename DWG Files",  use_container_width=True)
        st.markdown('<div class="btn-orange">', unsafe_allow_html=True)
        run_pdf = st.button("📄  Merge & Fix PDF Files", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with right:
        # Load mapping
        mapping = {}
        if excel_file:
            try:
                # Try multiple engines
                df_ex = None
                fname = excel_file.name.lower()
                for engine in (["xlrd"] if fname.endswith(".xls") else ["openpyxl","xlrd"]):
                    try:
                        excel_file.seek(0)
                        df_ex = pd.read_excel(excel_file, engine=engine, header=None)
                        break
                    except Exception:
                        continue

                if df_ex is None:
                    st.error("Cannot read Excel file — try saving as .xlsx format")
                else:
                    # Find header row — search broadly
                    hr = None
                    for i in range(min(30, len(df_ex))):
                        vals = [str(v).strip().lower() for v in df_ex.iloc[i] if str(v) not in("nan","None","")]
                        combined = " ".join(vals)
                        if any(x in combined for x in ["line number","line no","lineno"]):
                            hr = i; break
                        # fallback: look for row with many non-empty cells (likely header)
                    if hr is None:
                        # Try first row with more than 3 non-empty values
                        for i in range(min(10, len(df_ex))):
                            vals = [v for v in df_ex.iloc[i] if str(v) not in("nan","None","")]
                            if len(vals) >= 3:
                                hr = i; break

                    if hr is not None:
                        excel_file.seek(0)
                        df2 = pd.read_excel(excel_file, engine=engine, header=hr)
                        df2.columns = [str(c).strip().lower() for c in df2.columns]

                        # Show columns found for debugging
                        with st.expander("📋 Columns found in Excel"):
                            st.write(list(df2.columns))

                        # Find line number column
                        ln_col = next((c for c in df2.columns
                                       if any(x in c for x in ["line number","line no","lineno","tag","line_no"])), None)
                        # Find DWG name column
                        dw_col = next((c for c in df2.columns
                                       if any(x in c for x in ["dwg","drawing","doc"])), None)

                        if ln_col and dw_col:
                            for _, row in df2.iterrows():
                                ln = str(row[ln_col]).strip()
                                dw = str(row[dw_col]).strip()
                                if ln not in ("nan","None","") and dw not in ("nan","None",""):
                                    mapping[ln] = dw
                            st.success(f"✓ Excel loaded: {len(mapping)} line number → DWG mappings")
                        else:
                            st.warning(f"Could not find 'Line Number' or 'DWG' columns.\nFound: {list(df2.columns)}\nRename your columns to contain 'Line Number' and 'DWG'.")
                    else:
                        st.warning("Could not find header row in Excel file")
            except Exception as e:
                st.error(f"Excel error: {e}")

        # DWG Rename
        if run_dwg:
            if not dwg_files:
                st.warning("Upload DWG files first")
            elif not mapping:
                st.warning("Upload and load Excel sheet first")
            else:
                results = []
                for uf in dwg_files:
                    stem = Path(uf.name).stem
                    sh_match = re.search(r'\s+(SH[-_]?\d+)\s*$', stem, re.I)
                    sh_sfx = f" {sh_match.group(1).upper()}" if sh_match else ""
                    base = stem[:sh_match.start()].strip() if sh_match else stem.strip()
                    dwg_name = None
                    for lk,dn in mapping.items():
                        if lk.upper()==base.upper() or base.upper() in lk.upper() or lk.upper() in base.upper():
                            dwg_name=dn; break
                    if dwg_name:
                        new_name=f"{dwg_name}{sh_sfx}{dwg_suffix}.dwg"
                        results.append({"Original":uf.name,"Line No.":base,"New Name":new_name,"Status":"✓ OK"})
                    else:
                        results.append({"Original":uf.name,"Line No.":base,"New Name":"—","Status":"⚠ Not found"})
                st.session_state.dwg_results = results

        if st.session_state.get("dwg_results"):
            st.markdown("<div class='section-hdr'>DWG Rename Preview</div>", unsafe_allow_html=True)
            st.dataframe(pd.DataFrame(st.session_state.dwg_results),
                          use_container_width=True, hide_index=True)

        # PDF Merge
        if run_pdf:
            if not pdf_files:
                st.warning("Upload PDF files first")
            else:
                try:
                    import pypdf
                except ImportError:
                    st.error("pypdf not installed — run: pip install pypdf")
                    return

                groups = defaultdict(list)
                for uf in pdf_files:
                    stem = Path(uf.name).stem
                    sh_match = re.search(r'\s+(SH[-_]?\d+)\s*$', stem, re.I)
                    base = stem[:sh_match.start()].strip() if sh_match else stem.strip()
                    sh_num = int(re.search(r'\d+$', sh_match.group(1)).group()) if sh_match else 0
                    groups[base].append((sh_num, uf))

                pdf_results = []
                zip_buf = io.BytesIO()
                import zipfile
                with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for base, sheets in sorted(groups.items()):
                        dwg_name = None
                        if mapping:
                            for lk,dn in mapping.items():
                                if lk.upper()==base.upper() or base.upper() in lk.upper() or lk.upper() in base.upper():
                                    dwg_name=dn; break
                        out_name = f"{dwg_name or base}{pdf_suffix}.pdf"
                        sheets_sorted = [uf for _,uf in sorted(sheets)]
                        try:
                            writer = pypdf.PdfWriter()
                            for uf in sheets_sorted:
                                reader = pypdf.PdfReader(uf)
                                for page in reader.pages:
                                    w=float(page.mediabox.width); h=float(page.mediabox.height)
                                    if h>w: page.rotate(90)
                                    writer.add_page(page)
                            pdf_out = io.BytesIO()
                            writer.write(pdf_out)
                            zf.writestr(out_name, pdf_out.getvalue())
                            pdf_results.append({"Line No.":base,"Sheets":len(sheets_sorted),"Output PDF":out_name,"Status":"✓ OK"})
                        except Exception as e:
                            pdf_results.append({"Line No.":base,"Sheets":len(sheets_sorted),"Output PDF":out_name,"Status":f"✗ {e}"})

                st.session_state.pdf_results = pdf_results
                zip_buf.seek(0)
                st.download_button("⬇  Download Merged PDFs (ZIP)",
                                    data=zip_buf.getvalue(),
                                    file_name=f"PDFs_Merged_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
                                    mime="application/zip",
                                    use_container_width=True)

        if st.session_state.get("pdf_results"):
            st.markdown("<div class='section-hdr'>PDF Merge Preview</div>", unsafe_allow_html=True)
            st.dataframe(pd.DataFrame(st.session_state.pdf_results),
                          use_container_width=True, hide_index=True)

    st.markdown("<div class='footer'>Designed by / Eng. Shady Ahmed Elkenany</div>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
#  ROUTER
# ══════════════════════════════════════════════════════════════════
if not st.session_state.logged_in:
    page_login()
else:
    p = st.session_state.page
    if p == "launcher":    page_launcher()
    elif p == "insulation": page_insulation()
    elif p == "renamer":    page_renamer()
    else:                   page_launcher()
